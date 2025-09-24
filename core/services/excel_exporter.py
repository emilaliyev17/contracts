import logging
from datetime import datetime, date
from decimal import Decimal
from typing import List, Dict, Any
from django.db.models import QuerySet
from django.db import models
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from collections import defaultdict

from core.models import Contract, PaymentMilestone

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Service for exporting contract data to Excel format."""
    
    def __init__(self):
        self.workbook = None
        self.colors = {
            'header': 'FF4472C4',  # Blue
            'paid': 'FF90EE90',    # Light green
            'pending': 'FFFFFACD', # Light yellow
            'overdue': 'FFFFB6C1', # Light red
            'total': 'FFD3D3D3',   # Light gray
            'border': 'FF000000'   # Black
        }
    
    def export_contracts_to_excel(self, contracts_queryset: QuerySet, filename: str = None) -> str:
        """
        Export contracts and payment milestones to Excel file.
        
        Args:
            contracts_queryset: Django QuerySet of Contract objects
            filename: Optional filename (will generate if not provided)
            
        Returns:
            Path to the generated Excel file
        """
        try:
            logger.info(f"Starting Excel export for {contracts_queryset.count()} contracts")
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"contracts_export_{timestamp}.xlsx"
            
            logger.info(f"Using filename: {filename}")
            
            # Create workbook and worksheets
            logger.info("Creating workbook and worksheets")
            self.workbook = Workbook()
            self._setup_worksheets()
            logger.info("Worksheets created successfully")
            
            # Export data to each sheet
            logger.info("Exporting summary sheet")
            self._export_summary_sheet(contracts_queryset)
            logger.info("Summary sheet exported successfully")
            
            logger.info("Exporting contracts sheet")
            self._export_contracts_sheet(contracts_queryset)
            logger.info("Contracts sheet exported successfully")
            
            logger.info("Exporting payment schedule sheet")
            self._export_payment_schedule_sheet(contracts_queryset)
            logger.info("Payment schedule sheet exported successfully")
            
            # Save workbook
            logger.info(f"Saving workbook to {filename}")
            self.workbook.save(filename)
            logger.info(f"Workbook saved successfully to {filename}")
            
            logger.info(f"Excel export completed successfully: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}", exc_info=True)
            raise Exception(f"Excel export failed: {str(e)}")
    
    def _setup_worksheets(self):
        """Setup the Excel worksheets."""
        # Remove default sheet and create custom ones
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Create worksheets - Contracts sheet FIRST as main sheet
        self.contracts_sheet = self.workbook.create_sheet("Contracts", 0)
        self.payment_sheet = self.workbook.create_sheet("Payment Schedule", 1)
        self.summary_sheet = self.workbook.create_sheet("Summary", 2)
    
    def _export_summary_sheet(self, contracts_queryset: QuerySet):
        """Export simplified summary statistics to Summary sheet."""
        sheet = self.summary_sheet
        
        # Title
        sheet['A1'] = "Contract Export Summary"
        sheet['A1'].font = Font(size=16, bold=True, color="FF000080")
        sheet.merge_cells('A1:C1')
        
        # Generate summary data - SIMPLIFIED
        total_contracts = contracts_queryset.count()
        total_value = sum(c.total_value or 0 for c in contracts_queryset)
        total_milestones = PaymentMilestone.objects.filter(contract__in=contracts_queryset).count()
        
        # Get date range
        date_range_start = min((c.start_date for c in contracts_queryset if c.start_date), default=None)
        date_range_end = max((c.end_date for c in contracts_queryset if c.end_date), default=None)
        
        # Summary data - SIMPLIFIED
        summary_data = [
            ["Metric", "Value", ""],
            ["Total Contracts", total_contracts, ""],
            ["Total Contract Value", f"${total_value:,.2f}", "USD"],
            ["Total Payment Milestones", total_milestones, ""],
            ["", "", ""],
            ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M"), ""],
            ["Date Range", f"{date_range_start} to {date_range_end}" if date_range_start else "N/A", ""],
        ]
        
        # Write summary data
        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Format header row
                if row_idx == 3:
                    cell.font = Font(bold=True, color="FFFFFFFF")
                    cell.fill = PatternFill(start_color=self.colors['header'], 
                                          end_color=self.colors['header'], 
                                          fill_type="solid")
                # Format amount columns
                elif col_idx == 3 and value != "":
                    cell.number_format = '$#,##0.00'
                
                # Add borders
                cell.border = self._get_border()
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _export_contracts_sheet(self, contracts_queryset: QuerySet):
        """Export ALL contract details to Contracts sheet - MAIN SHEET."""
        sheet = self.contracts_sheet
        
        # Add title
        sheet['A1'] = f"Contract Details - {contracts_queryset.count()} Contracts"
        sheet['A1'].font = Font(size=14, bold=True, color="FF000080")
        sheet.merge_cells('A1:I1')
        
        # Headers - Focused on what users need
        headers = [
            "Contract Name",
            "Client",
            "Contract Number",
            "Total Value",
            "Currency",
            "Start Date",
            "End Date",
            "Status",
            "# of Milestones"
        ]
        
        # Write headers (row 3 to leave space for title)
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill(start_color=self.colors['header'], 
                                  end_color=self.colors['header'], 
                                  fill_type="solid")
            cell.border = self._get_border()
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write ALL contract data - no filtering
        row_start = 4  # Start after header
        for row_idx, contract in enumerate(contracts_queryset, start=row_start):
            try:
                logger.debug(f"Processing contract {contract.id}: {contract.contract_name}")
                
                milestone_count = contract.payment_milestones.count()
                
                # Handle null dates safely
                try:
                    start_date = contract.start_date.strftime('%m/%d/%Y') if contract.start_date else ""
                except Exception as e:
                    logger.warning(f"Start date formatting error for contract {contract.id}: {e}")
                    start_date = ""
                
                try:
                    end_date = contract.end_date.strftime('%m/%d/%Y') if contract.end_date else ""
                except Exception as e:
                    logger.warning(f"End date formatting error for contract {contract.id}: {e}")
                    end_date = ""
                
                try:
                    upload_date = contract.upload_date.strftime("%Y-%m-%d %H:%M") if contract.upload_date else ""
                except Exception as e:
                    logger.warning(f"Upload date formatting error for contract {contract.id}: {e}")
                    upload_date = ""
                
                # Safely get all contract fields with null checks - matching new headers
                row_data = [
                    contract.contract_name or "N/A",
                    contract.client_name or "N/A",
                    contract.contract_number or "N/A",
                    contract.total_value or 0,
                    contract.currency or "USD",
                    start_date or "N/A",
                    end_date or "Ongoing",
                    contract.status or "unknown",
                    milestone_count
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    try:
                        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Format currency - Column 4 is Total Value
                        if col_idx == 4:  
                            if value and value != 0:
                                cell.number_format = '$#,##0.00'
                            else:
                                cell.value = "N/A"
                        
                        # Format dates - Columns 6 and 7
                        elif col_idx in [6, 7]:
                            cell.alignment = Alignment(horizontal='center')
                        
                        # Format status with color coding - Column 8
                        elif col_idx == 8:  
                            cell.alignment = Alignment(horizontal='center')
                            if value == 'completed':
                                cell.fill = PatternFill(start_color=self.colors['paid'], 
                                                      end_color=self.colors['paid'], 
                                                      fill_type="solid")
                            elif value == 'processing':
                                cell.fill = PatternFill(start_color=self.colors['pending'], 
                                                      end_color=self.colors['pending'], 
                                                      fill_type="solid")
                            elif value in ['error', 'failed']:
                                cell.fill = PatternFill(start_color=self.colors['overdue'], 
                                                      end_color=self.colors['overdue'], 
                                                      fill_type="solid")
                        
                        # Center align milestone count - Column 9
                        elif col_idx == 9:
                            cell.alignment = Alignment(horizontal='center')
                        
                        cell.border = self._get_border()
                        
                    except Exception as e:
                        logger.error(f"Error writing cell {col_idx} for contract {contract.id}: {e}")
                        # Write error message to cell
                        sheet.cell(row=row_idx, column=col_idx, value=f"ERROR: {str(e)}")
                        
            except Exception as e:
                logger.error(f"Error processing contract {contract.id}: {e}")
                # Write error row with proper columns
                error_row = [f"ERROR - Contract {contract.id}", "", "", 0, "", "", "", "error", 0]
                for col_idx, value in enumerate(error_row, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = PatternFill(start_color=self.colors['overdue'], 
                                          end_color=self.colors['overdue'], 
                                          fill_type="solid")
                continue
        
        # Add totals row
        total_row = row_start + contracts_queryset.count()
        sheet.cell(row=total_row, column=1, value="TOTAL")
        sheet.cell(row=total_row, column=1).font = Font(bold=True)
        sheet.cell(row=total_row, column=4, value=f"=SUM(D{row_start}:D{total_row-1})")
        sheet.cell(row=total_row, column=4).number_format = '$#,##0.00'
        sheet.cell(row=total_row, column=4).font = Font(bold=True)
        sheet.cell(row=total_row, column=4).fill = PatternFill(start_color=self.colors['total'], 
                                                            end_color=self.colors['total'], 
                                                            fill_type="solid")
        sheet.cell(row=total_row, column=9, value=f"=SUM(I{row_start}:I{total_row-1})")
        sheet.cell(row=total_row, column=9).font = Font(bold=True)
        sheet.cell(row=total_row, column=9).alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _export_payment_schedule_sheet(self, contracts_queryset: QuerySet):
        """Export payment milestones to Payment Schedule sheet organized by contract."""
        sheet = self.payment_sheet
        
        # Add title
        sheet['A1'] = f"Payment Schedule - All Milestones"
        sheet['A1'].font = Font(size=14, bold=True, color="FF000080")
        sheet.merge_cells('A1:F1')
        
        # Headers - Simplified
        headers = [
            "Contract Name",
            "Milestone Name",
            "Due Date",
            "Amount",
            "Status"
        ]
        
        # Write headers (row 3 to leave space for title)
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill(start_color=self.colors['header'], 
                                  end_color=self.colors['header'], 
                                  fill_type="solid")
            cell.border = self._get_border()
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get all payment milestones for these contracts - order by contract then date
        milestones = PaymentMilestone.objects.filter(
            contract__in=contracts_queryset
        ).select_related('contract').order_by('contract__contract_name', 'due_date')
        
        # Write milestone data
        row_start = 4
        for row_idx, milestone in enumerate(milestones, start=row_start):
            try:
                logger.debug(f"Processing milestone {milestone.id} for contract {milestone.contract.id}")
                
                # Handle null dates safely
                try:
                    due_date = milestone.due_date.strftime('%m/%d/%Y') if milestone.due_date else ""
                except Exception as e:
                    logger.warning(f"Date formatting error for milestone {milestone.id}: {e}")
                    due_date = ""
                
                # Safely get all milestone fields with null checks - matching new headers
                row_data = [
                    milestone.contract.contract_name if milestone.contract else "N/A",
                    milestone.milestone_name or "Payment",
                    due_date or "TBD",
                    milestone.amount or 0,
                    milestone.status or "pending"
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    try:
                        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Format currency - Column 4 is Amount
                        if col_idx == 4:  
                            if value and value != 0:
                                cell.number_format = '$#,##0.00'
                            else:
                                cell.value = "N/A"
                        
                        # Format date - Column 3
                        elif col_idx == 3:
                            cell.alignment = Alignment(horizontal='center')
                        
                        # Color code status - Column 5
                        elif col_idx == 5:
                            cell.alignment = Alignment(horizontal='center')
                            if value == 'paid':
                                cell.fill = PatternFill(start_color=self.colors['paid'], 
                                                      end_color=self.colors['paid'], 
                                                      fill_type="solid")
                            elif value == 'pending':
                                cell.fill = PatternFill(start_color=self.colors['pending'], 
                                                      end_color=self.colors['pending'], 
                                                      fill_type="solid")
                            elif value == 'overdue':
                                cell.fill = PatternFill(start_color=self.colors['overdue'], 
                                                      end_color=self.colors['overdue'], 
                                                      fill_type="solid")
                        
                        cell.border = self._get_border()
                        
                    except Exception as e:
                        logger.error(f"Error writing cell {col_idx} for milestone {milestone.id}: {e}")
                        # Write error message to cell
                        sheet.cell(row=row_idx, column=col_idx, value=f"ERROR: {str(e)}")
                        
            except Exception as e:
                logger.error(f"Error processing milestone {milestone.id}: {e}")
                # Write error row with proper columns
                error_row = [f"ERROR", f"Milestone {milestone.id}", "", 0, "error"]
                for col_idx, value in enumerate(error_row, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = PatternFill(start_color=self.colors['overdue'], 
                                          end_color=self.colors['overdue'], 
                                          fill_type="solid")
                continue
        
        # Add totals row if there are milestones
        if milestones.exists():
            total_row = row_start + milestones.count()
            sheet.cell(row=total_row, column=1, value="TOTAL")
            sheet.cell(row=total_row, column=1).font = Font(bold=True)
            sheet.cell(row=total_row, column=4, value=f"=SUM(D{row_start}:D{total_row-1})")
            sheet.cell(row=total_row, column=4).number_format = '$#,##0.00'
            sheet.cell(row=total_row, column=4).font = Font(bold=True)
            sheet.cell(row=total_row, column=4).fill = PatternFill(start_color=self.colors['total'], 
                                                                end_color=self.colors['total'], 
                                                                fill_type="solid")
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _get_border(self):
        """Get a standard border for cells."""
        thin_border = Border(
            left=Side(style='thin', color=self.colors['border']),
            right=Side(style='thin', color=self.colors['border']),
            top=Side(style='thin', color=self.colors['border']),
            bottom=Side(style='thin', color=self.colors['border'])
        )
        return thin_border
